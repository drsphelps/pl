from openpyxl import load_workbook
from dbManager import addArtist, addAlbum, addSong

class extraWorkbook():
    def __init__(self, file):
        self.__workbook = load_workbook(filename=file)

    # Allows us to read a sheet, given a number rather than a name
    def getSheet(self, num):
        sheet_title = self.__workbook.sheetnames[num]
        return extraSheet(self.__workbook[sheet_title])

class extraSheet():
    def __init__(self, s):
        self.emptySheet = False
        self.__sheet = s
        # After reading header we need to start at row 2 to get the rest of the data
        self.__rowsRead = 2
        self.__headers = []
        self.__readHeaders()
        self.__rows = []
        if len(self.__headers) != 0:
        # Stores a list of rows each of which is a list containing the info for that row
            self.readAllRows()
        else:
            emptySheet = True
        self.addToDatabase()

    def addToDatabase(self):
        artistsAdded = []
        albumsAdded = []
        songsAdded = []
        for row in self.__rows:
            print(row)
            if row['Band'] is not None:
                if row['Band'] not in artistsAdded:
                    addArtist(row['Band'])
                    artistsAdded.append(row['Band'])
                if row['Album'] not in albumsAdded:
                    addAlbum(row['Album'], row['Bandcamp Link'], row['Band'])
                    albumsAdded.append(row['Album'])
                if row['Playable'] is None:
                    row['Playable'] = 1
                else: row['Playable'] = 0
                if row['Play Count'] is None:
                    row['Play Count'] = 0
                addSong(row['Song'], row['Length'], row['Band'], row['Album'], row['Playable'], row['Play Count'], row['iTunes Link'])
                songsAdded.append(row['Song'])

    def getRowsRead(self):
        return self.__rowsRead

    def setRowsRead(self, num):
        self.__rowsRead = num

    def __readHeaders(self):
        for col in range(65, 91):
            cell = chr(col) + "1"
            value = self.__sheet[cell].value
            if value is not None:
                self.__headers.append(value)

    # Read every row in the file
    def readAllRows(self):
        for row in range(2, self.__sheet.max_row):
            rowToAdd = {}
            for col in range(0, len(self.__headers)):
                cell = chr(65+col)+str(row)
                rowToAdd[self.__headers[col]]=self.__sheet[cell].value
            if rowToAdd['Length'] is not None:
                rowToAdd['Length'] = rowToAdd['Length'].strftime("%H:%M:%S")
            self.__rows.append(rowToAdd)
        self.__rowsRead = self.__sheet.max_row

    # Read the rows following the ones we have already read
    def readFromEnd(self):
        for row in range(self.__rowsRead, self.__sheet.max_row):
            rowToAdd = []
            for col in range(65, 65+len(self.__headers)):
                cell = chr(col)+str(row)
                rowToAdd.append(self.__sheet[cell].value)
            self.__rows.append(rowToAdd)
        print(len(self.__rows))

if __name__ == "__main__":
    workbook = "SkaAndReggaeDatabase.xlsx"
    w = extraWorkbook(workbook)
    s = w.getSheet(0)
